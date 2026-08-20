import csv
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from openpyxl import load_workbook
from app.database import get_db
from app import models, schemas, utils

router = APIRouter(prefix="/api/students", tags=["students"])


@router.get("/", response_model=list[schemas.StudentRead])
async def list_students(
    section_id: Optional[int] = Query(None),
    academic_year_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    q = select(models.Student)
    if section_id is not None:
        q = q.where(models.Student.section_id == section_id)
    if academic_year_id is not None:
        q = q.where(models.Student.academic_year_id == academic_year_id)
    result = await db.execute(q.order_by(models.Student.roll_no))
    return result.scalars().all()


@router.get("/{student_id}", response_model=schemas.StudentRead)
async def get_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    result = await db.execute(select(models.Student).where(models.Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.post("/", response_model=schemas.StudentRead)
async def create_student(
    payload: schemas.StudentCreate,
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    result = await db.execute(select(models.Section).where(models.Section.id == payload.section_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Section not found")
    result = await db.execute(
        select(models.AcademicYear).where(models.AcademicYear.id == payload.academic_year_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Academic year not found")

    student = models.Student(**payload.model_dump())
    db.add(student)
    await db.commit()
    await db.refresh(student)
    await utils.audit_log(db, user.id, "create", "student", student.id, payload.model_dump())
    return student


@router.put("/{student_id}", response_model=schemas.StudentRead)
async def update_student(
    student_id: int,
    payload: schemas.StudentCreate,
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    result = await db.execute(select(models.Student).where(models.Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    old_data = {c.name: getattr(student, c.name) for c in student.__table__.columns}
    for k, v in payload.model_dump().items():
        setattr(student, k, v)
    await db.commit()
    await db.refresh(student)
    await utils.audit_log(db, user.id, "update", "student", student.id, {"old": old_data, "new": payload.model_dump()})
    return student


@router.delete("/{student_id}")
async def delete_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    result = await db.execute(select(models.Student).where(models.Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    await db.delete(student)
    await db.commit()
    await utils.audit_log(db, user.id, "delete", "student", student_id, {"roll_no": student.roll_no, "name": student.name})
    return {"detail": "Student deleted"}


@router.post("/import/preview", response_model=schemas.StudentImportPreview)
async def import_preview(
    section_id: int = Query(...),
    academic_year_id: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    content = await file.read()
    rows: list[dict] = []
    errors: list[dict] = []

    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):
            rows.append(dict(row))
    elif filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            d = {}
            for h, v in zip(headers, row):
                d[h] = str(v).strip() if v is not None else ""
            rows.append(d)
        wb.close()
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .csv or .xlsx")

    valid: list[schemas.StudentBulkItem] = []
    for i, row in enumerate(rows, start=2):
        roll = row.get("roll_no", row.get("roll number", row.get("rollno", ""))).strip()
        name = row.get("name", row.get("student_name", "")).strip()
        email = row.get("email", "").strip() or None
        errs = []
        if not roll:
            errs.append("Missing roll_no")
        if not name:
            errs.append("Missing name")
        if errs:
            errors.append({"row": i, "errors": errs})
        else:
            valid.append(schemas.StudentBulkItem(roll_no=roll, name=name, email=email))

    return schemas.StudentImportPreview(total_rows=len(rows), valid_rows=valid, errors=errors)


@router.post("/import/confirm")
async def import_confirm(
    section_id: int = Query(...),
    academic_year_id: int = Query(...),
    students: list[schemas.StudentBulkItem] = [],
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    result = await db.execute(select(models.Section).where(models.Section.id == section_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Section not found")
    result = await db.execute(
        select(models.AcademicYear).where(models.AcademicYear.id == academic_year_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Academic year not found")

    created = 0
    for s in students:
        student = models.Student(
            roll_no=s.roll_no,
            name=s.name,
            email=s.email,
            section_id=section_id,
            academic_year_id=academic_year_id,
        )
        db.add(student)
        created += 1
    await db.commit()
    await utils.audit_log(db, user.id, "import", "student", None, {"count": created, "section_id": section_id})
    return {"detail": f"Imported {created} students"}


class BulkImportRequest(BaseModel):
    students: list[schemas.StudentBulkItem]
    section_id: int
    academic_year_id: int


@router.post("/import", response_model=list[schemas.StudentRead])
async def bulk_import(
    payload: BulkImportRequest,
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    result = await db.execute(select(models.Section).where(models.Section.id == payload.section_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Section not found")
    result = await db.execute(
        select(models.AcademicYear).where(models.AcademicYear.id == payload.academic_year_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Academic year not found")

    created_students: list[models.Student] = []
    for s in payload.students:
        student = models.Student(
            roll_no=s.roll_no,
            name=s.name,
            email=s.email,
            section_id=payload.section_id,
            academic_year_id=payload.academic_year_id,
        )
        db.add(student)
        created_students.append(student)
    await db.commit()
    await utils.audit_log(db, user.id, "import", "student", None, {"count": len(created_students), "section_id": payload.section_id})
    return created_students
