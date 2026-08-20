from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.database import get_db
from app import models, utils

router = APIRouter(prefix="/api/export", tags=["export"])


class BackupRequest(BaseModel):
    data: dict


FILL_RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
FILL_GREEN = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@router.get("/excel")
async def export_excel(
    section_id: Optional[int] = Query(None),
    subject_id: Optional[int] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    students_result = await db.execute(
        select(models.Student)
        .where(models.Student.section_id == section_id if section_id else True)
        .order_by(models.Student.roll_no)
    )
    students = students_result.scalars().all()

    att_q = select(models.Attendance).order_by(models.Attendance.date)
    if section_id:
        student_ids = select(models.Student.id).where(models.Student.section_id == section_id)
        att_q = att_q.where(models.Attendance.student_id.in_(student_ids))
    if subject_id:
        att_q = att_q.where(models.Attendance.subject_id == subject_id)
    if start_date:
        att_q = att_q.where(models.Attendance.date >= start_date)
    if end_date:
        att_q = att_q.where(models.Attendance.date <= end_date)

    records_result = await db.execute(att_q)
    records = records_result.scalars().all()

    subject_result = await db.execute(select(models.Subject).where(models.Subject.id == subject_id)) if subject_id else None
    subject = subject_result.scalar_one_or_none() if subject_result else None
    section_result = await db.execute(select(models.Section).where(models.Section.id == section_id)) if section_id else None
    section = section_result.scalar_one_or_none() if section_result else None

    wb = Workbook()

    # ── Sheet 1: Raw Data ──
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    raw_headers = ["Student ID", "Roll No", "Student Name", "Date", "Status"]
    for col, header in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    student_map = {s.id: s for s in students}
    for row_idx, rec in enumerate(records, 2):
        stu = student_map.get(rec.student_id)
        vals = [
            rec.student_id,
            stu.roll_no if stu else "",
            stu.name if stu else "",
            rec.date.isoformat(),
            rec.status.value,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws_raw.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(raw_headers) + 1):
        ws_raw.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 2: Summary with formulas ──
    ws_sum = wb.create_sheet("Summary")
    sum_headers = [
        "Roll No", "Student Name", "Total Sessions",
        "Present", "Absent", "Late", "Excused", "Attendance %"
    ]
    for col, header in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row_idx, stu in enumerate(students, 2):
        ws_sum.cell(row=row_idx, column=1, value=stu.roll_no).border = BORDER
        ws_sum.cell(row=row_idx, column=2, value=stu.name).border = BORDER

        stu_row_range = f"Raw Data!$E$2:$E$10000"

        # Total sessions = count of all records for this student
        total_formula = f'=COUNTIF(Raw Data!$A$2:$A$10000,A{row_idx})'
        ws_sum.cell(row=row_idx, column=3, value=total_formula).border = BORDER

        # Present count
        present_formula = (
            f'=COUNTIFS({stu_row_range},"present",'
            f'Raw Data!$A$2:$A$10000,A{row_idx})'
        )
        ws_sum.cell(row=row_idx, column=4, value=present_formula).border = BORDER

        # Absent count
        absent_formula = (
            f'=COUNTIFS({stu_row_range},"absent",'
            f'Raw Data!$A$2:$A$10000,A{row_idx})'
        )
        ws_sum.cell(row=row_idx, column=5, value=absent_formula).border = BORDER

        # Late count
        late_formula = (
            f'=COUNTIFS({stu_row_range},"late",'
            f'Raw Data!$A$2:$A$10000,A{row_idx})'
        )
        ws_sum.cell(row=row_idx, column=6, value=late_formula).border = BORDER

        # Excused count
        excused_formula = (
            f'=COUNTIFS({stu_row_range},"excused",'
            f'Raw Data!$A$2:$A$10000,A{row_idx})'
        )
        ws_sum.cell(row=row_idx, column=7, value=excused_formula).border = BORDER

        # Percentage = (present + late) / total * 100
        pct_formula = f'=IF(C{row_idx}>0,((D{row_idx}+F{row_idx})/C{row_idx})*100,0)'
        pct_cell = ws_sum.cell(row=row_idx, column=8, value=pct_formula)
        pct_cell.border = BORDER
        pct_cell.number_format = "0.00"

        # Conditional formatting on percentage column
        for col_idx in range(1, len(sum_headers) + 1):
            c = ws_sum.cell(row=row_idx, column=col_idx)
            c.alignment = Alignment(horizontal="center")

    # Apply conditional formatting rules manually via openpyxl
    from openpyxl.formatting.rule import CellIsRule
    last_row = len(students) + 1
    pct_range = f"H2:H{last_row}"
    ws_sum.conditional_formatting.add(pct_range, CellIsRule(
        operator="lessThan", formula=["75"], fill=FILL_RED
    ))
    ws_sum.conditional_formatting.add(pct_range, CellIsRule(
        operator="between", formula=["75", "85"], fill=FILL_YELLOW
    ))
    ws_sum.conditional_formatting.add(pct_range, CellIsRule(
        operator="greaterThan", formula=["85"], fill=FILL_GREEN
    ))

    for col in range(1, len(sum_headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    filename = f"attendance_{subject.code if subject else 'export'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/backup")
async def backup_data(
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    def model_to_dict(m):
        return {c.name: getattr(m, c.name) for c in m.__table__.columns}

    entities = {}
    for name, model in [
        ("students", models.Student),
        ("sections", models.Section),
        ("subjects", models.Subject),
        ("attendance", models.Attendance),
        ("academic_years", models.AcademicYear),
    ]:
        rows = (await db.execute(select(model))).scalars().all()
        entities[name] = [model_to_dict(r) for r in rows]

    import json
    data_bytes = json.dumps({"backup": True, "entities": entities}, default=str, indent=2).encode()
    buf = BytesIO(data_bytes)
    return StreamingResponse(
        buf,
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=markit-backup.json"},
    )


@router.post("/restore")
async def restore_data(
    payload: BackupRequest,
    db: AsyncSession = Depends(get_db),
    user: models.User = Depends(utils.get_current_user),
):
    data = payload.data
    if not data.get("backup"):
        raise HTTPException(status_code=400, detail="Invalid backup format")
    entities = data.get("entities", {})
    table_map = {
        "students": models.Student.__table__,
        "sections": models.Section.__table__,
        "subjects": models.Subject.__table__,
        "attendance": models.Attendance.__table__,
        "academic_years": models.AcademicYear.__table__,
    }
    for table in table_map.values():
        await db.execute(table.delete())
    for entity_name, rows in entities.items():
        table = table_map.get(entity_name)
        if table is None:
            continue
        for row in rows:
            filtered = {k: v for k, v in row.items() if k in table.columns}
            await db.execute(table.insert().values(**filtered))
    await db.commit()
    await utils.audit_log(db, user.id, "restore", "backup", None, {"source": "manual"})
    return {"detail": "Data restored successfully"}
